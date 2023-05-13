'''
A program specifically designed for monitoring the status of EP patents.
Author: Lancer1911 2021.09.23
'''

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles.alignment import Alignment
import os
import tkinter as tk 
from tkinter import filedialog
from os import path
from tkinter import messagebox
from tkinter import simpledialog
import sys
import platform
import time

root = tk.Tk()
root.withdraw()

user_confirm = messagebox.askokcancel("Notice","You are using a script to scrape EP patent and patent application status from register.epo.org.\n\nPlease make sure you know the following:\n1. The location of the Excel file to be processed;\n2. The column where the EP patent numbers are located, such as column C.")
if user_confirm ==  False:
    sys.exit()

def specify_file(initial_directory, title, file_types):
    file_path = None
    while not file_path:
       file_path= filedialog.askopenfilename(initialdir = initial_directory, title = title, filetypes = file_types)
       if file_path:
           return file_path

current_directory = path.dirname(__file__)
file_path = specify_file(path.join(current_directory, '%USERPROFILE%/Downloads'), 'Select the Excel file with EP patent numbers', (('Excel files','*.xlsx'),('all files','*.*')))
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

target_column = None
while not target_column:
    user_input = simpledialog.askstring("Input", "Input the column where the EP patent numbers are located, e.g., C:",parent=root)
    try:
        target_column = ord(user_input)
        if target_column >=65 and target_column <=90:
            target_column = target_column - 64
        elif target_column >=97 and target_column <=122:
            target_column = target_column - 96
        else:
            messagebox.showerror("Error", "Please input a letter!")
            target_column = None
    except ValueError:
        messagebox.showerror("Error", "Please input a letter!")
        target_column = None
    except TypeError:
        sys.exit()

target_row = None
while not target_row:
    user_input = simpledialog.askstring("Input", "From which row to start scraping?\nNote: You must input 2 the first time you use it, otherwise it will overwrite the original spreadsheet column.",parent=root)
    try:
        target_row = int(user_input)
    except ValueError:
        messagebox.showerror("Error", "Please input an integer!")
    except TypeError:
        sys.exit()

if target_row <= 2:
    sheet.insert_cols(target_column + 1)
    target_row = 2

sheet.column_dimensions[chr(ord('@') + target_column + 1)].width = 27.5

max_row = sheet.max_row

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--ignore-certificate-errors')
chrome_options.add_argument("--test-type")

# Get the current directory
current_dir = os.path.dirname(os.path.realpath(__file__))

# Concatenate the path to chromedriver
chromedriver_path = os.path.join(current_dir, 'chromedriver')

# Initialize webdriver with chromedriver path
driver = webdriver.Chrome(executable_path=chromedriver_path, options=chrome_options)

print("\n\033[1m\033[36m==================================\n  Start to Scrape Patent Status!\n==================================\033[0m")

for i in range(target_row, max_row + 1): 
    cell_obj = sheet.cell(row = i, column = target_column)
    cell_target = sheet.cell(row = i, column = target_column + 1)
    print("\n\033[37m\033[42m", i, ": \033[0m\033[32m", cell_obj.value, "\033[0m")
    time.sleep(5)
    URL = str("https://register.epo.org")
    driver.get(URL)

    search_bar = driver.find_element_by_xpath("//*[@name='query']")
    search_bar.clear()
    search_bar.send_keys(cell_obj.value)
    driver.find_element_by_xpath("//*[@type='submit']").click()

    status = "null"
    try:
        table_data_elements = driver.find_elements_by_xpath("//td")
        for table_data in table_data_elements :
            if table_data.text == "Status":
                status = table_data.find_element_by_xpath("./following-sibling::td").text
    except:
        status = "not found"

    print(status)
    cell_target.value = status
    cell_target.alignment = Alignment(wrap_text=True)

    if i % 10 == 0:
        workbook.save(file_path)
        print("\n\033[33m=========================================\nThe status of item No." , i, "has been saved.\n=========================================\033[0m")

print("\n\033[1m\033[36m==================================\nFinished! Hope you enjoy it...\n==================================\033[0m")

cell_target = sheet.cell(1,target_column + 1)
cell_target.value =  "Status"

workbook.save(file_path)

driver.close()

